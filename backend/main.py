"""
Dissertation Experiment Scorer - FastAPI Backend
"""
import os
import sys
import tempfile
import shutil
import json
from datetime import datetime, timedelta
from typing import Optional, List
from pathlib import Path

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import bcrypt
from jose import JWTError, jwt

# Add scripts path for qualtrics_cleanup
sys.path.insert(0, '/home/silver/clawd/scripts')
from qualtrics_cleanup import process_csv

# ============================================
# CONFIG
# ============================================

SECRET_KEY = os.environ.get("JWT_SECRET", "dissertation-scorer-secret-change-in-prod")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Upload history file
UPLOAD_HISTORY_FILE = Path(__file__).parent / "upload_history.json"

# Approved users (email -> hashed password)
# Password: admin
USERS = {
    "projectskyfall24@gmail.com": {
        "name": "Kai",
        "password_hash": bcrypt.hashpw(b"admin", bcrypt.gensalt()).decode()
    },
    "rafiamejri17@gmail.com": {
        "name": "Rafia",
        "password_hash": bcrypt.hashpw(b"admin", bcrypt.gensalt()).decode()
    }
}

# ============================================
# MODELS
# ============================================

class LoginRequest(BaseModel):
    email: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    name: str

class UserInfo(BaseModel):
    email: str
    name: str

# ============================================
# APP SETUP
# ============================================

app = FastAPI(title="Dissertation Scorer API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Will be restricted in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Processing-Stats"],  # Expose custom header to frontend
)

security = HTTPBearer()

# ============================================
# AUTH HELPERS
# ============================================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode(), hashed_password.encode())

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> UserInfo:
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None or email not in USERS:
            raise HTTPException(status_code=401, detail="Invalid token")
        return UserInfo(email=email, name=USERS[email]["name"])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============================================
# ROUTES
# ============================================

@app.get("/health")
async def health():
    return {"status": "ok", "service": "dissertation-scorer"}

@app.post("/auth/login", response_model=TokenResponse)
async def login(request: LoginRequest):
    email = request.email.lower().strip()
    
    if email not in USERS:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    user = USERS[email]
    if not verify_password(request.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token(data={"sub": email})
    return TokenResponse(access_token=access_token, name=user["name"])

@app.get("/auth/me", response_model=UserInfo)
async def get_me(user: UserInfo = Depends(get_current_user)):
    return user

@app.get("/uploads/history")
async def get_upload_history(user: UserInfo = Depends(get_current_user)):
    """Get upload history."""
    return load_upload_history()

class ProcessingStats(BaseModel):
    total_responses: int
    complete: int
    incomplete: int
    completion_rate: float
    conditions: dict
    mean_recall_score: float
    recall_score_range: dict
    gender_breakdown: dict

class UploadLogEntry(BaseModel):
    name: str
    timestamp: str
    responses: int
    complete: int

def load_upload_history() -> List[dict]:
    """Load upload history from file."""
    if UPLOAD_HISTORY_FILE.exists():
        try:
            return json.loads(UPLOAD_HISTORY_FILE.read_text())
        except:
            return []
    return []

def save_upload_log(name: str, responses: int, complete: int):
    """Save an upload log entry."""
    history = load_upload_history()
    history.append({
        "name": name,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "responses": responses,
        "complete": complete
    })
    # Keep only last 50 entries
    history = history[-50:]
    UPLOAD_HISTORY_FILE.write_text(json.dumps(history, indent=2))

@app.post("/process/csv")
async def process_csv_file(
    file: UploadFile = File(...),
    output_format: str = "xlsx",
    user: UserInfo = Depends(get_current_user)
):
    """Process uploaded CSV file and return cleaned/scored output with stats."""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Please upload a CSV file")
    
    # Create temp directory for processing
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / "input.csv"
        output_ext = "xlsx" if output_format == "xlsx" else "csv"
        output_path = Path(tmpdir) / f"scored_output.{output_ext}"
        
        # Save uploaded file
        with open(input_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Process the file
        try:
            process_csv(str(input_path), str(output_path))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")
        
        if not output_path.exists():
            raise HTTPException(status_code=500, detail="Processing failed - no output generated")
        
        # Calculate stats from output file
        stats = calculate_stats(str(output_path))
        
        # Log the upload
        save_upload_log(
            name=user.name,
            responses=stats.get("total_responses", 0),
            complete=stats.get("complete", 0)
        )
        
        # Read output file and return
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if output_ext == "xlsx" else "text/csv"
        
        # Copy to a persistent temp location for download
        # Use mkstemp to avoid file handle conflicts
        fd, final_output_path = tempfile.mkstemp(suffix=f".{output_ext}", prefix="scorer_")
        os.close(fd)  # Close the file descriptor immediately
        shutil.copy(output_path, final_output_path)
        
        # Return stats in headers and file as response
        response = FileResponse(
            final_output_path,
            media_type=media_type,
            filename=f"scored_results.{output_ext}"
        )
        # Add stats as custom header (JSON encoded)
        import json
        response.headers["X-Processing-Stats"] = json.dumps(stats)
        return response


def calculate_stats(output_path: str) -> dict:
    """Calculate summary statistics from processed output file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        # Helper to get column values
        def get_column_values(col_name):
            if col_name not in headers:
                return []
            idx = headers.index(col_name)
            return [row[idx].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row)]
        
        # Basic counts
        completion_status = get_column_values('CompletionStatus')
        total = len(completion_status)
        complete = sum(1 for s in completion_status if s == 'COMPLETE')
        incomplete = total - complete
        
        # Condition distribution
        conditions = {}
        for c in get_column_values('Condition'):
            if c:
                c_lower = str(c).lower().strip()
                conditions[c_lower] = conditions.get(c_lower, 0) + 1
        
        # Recall scores (only for complete entries)
        recall_scores = []
        recall_total_idx = headers.index('Recall_Total') if 'Recall_Total' in headers else None
        comp_idx = headers.index('CompletionStatus') if 'CompletionStatus' in headers else None
        
        if recall_total_idx is not None and comp_idx is not None:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[comp_idx].value == 'COMPLETE':
                    score = row[recall_total_idx].value
                    if isinstance(score, (int, float)):
                        recall_scores.append(score)
        
        mean_recall = sum(recall_scores) / len(recall_scores) if recall_scores else 0
        
        # Gender breakdown
        gender_breakdown = {}
        for g in get_column_values('Gender'):
            if g:
                g_str = str(g).strip()
                gender_breakdown[g_str] = gender_breakdown.get(g_str, 0) + 1
        
        # Age range
        ages = []
        for a in get_column_values('Age'):
            if a:
                try:
                    age_val = int(a)
                    if 10 < age_val < 120:  # Sanity check
                        ages.append(age_val)
                except (ValueError, TypeError):
                    pass
        
        return {
            "total_responses": total,
            "complete": complete,
            "incomplete": incomplete,
            "completion_rate": round(complete / total * 100, 1) if total > 0 else 0,
            "conditions": conditions,
            "mean_recall_score": round(mean_recall, 2),
            "recall_score_range": {
                "min": min(recall_scores) if recall_scores else 0,
                "max": max(recall_scores) if recall_scores else 0
            },
            "gender_breakdown": gender_breakdown,
            "age_range": {
                "min": min(ages) if ages else 0,
                "max": max(ages) if ages else 0
            }
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/help/export")
async def help_export(user: UserInfo = Depends(get_current_user)):
    """Return instructions for exporting from Qualtrics."""
    return {
        "title": "How to Export from Qualtrics",
        "steps": [
            "1. Log in to your Qualtrics account",
            "2. Open your survey project",
            "3. Click 'Data & Analysis' in the top navigation",
            "4. Click 'Export & Import' → 'Export Data'",
            "5. Select 'CSV' as the format",
            "6. Choose 'Download all fields' or select the fields you need",
            "7. Click 'Download' and save the file",
            "8. Upload the CSV file here for processing"
        ]
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
