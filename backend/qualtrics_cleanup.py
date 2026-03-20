#!/usr/bin/env python3
"""
Qualtrics CSV Cleanup & Recall Scoring Script
For Kai's regulatory focus × task framing memory experiment

Usage: python3 qualtrics_cleanup.py input.csv output.csv
       python3 qualtrics_cleanup.py input.csv output.xlsx  (for colored Excel)
"""

import sys
import csv
import re
from difflib import SequenceMatcher

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ============================================
# WORD LISTS BY BLOCK AND VALENCE
# ============================================

WORDS = {
    1: {
        'positive': ['laughter', 'cheer', 'affection', 'paradise', 'hug', 'kiss', 'comedy'],
        'neutral': ['bathroom', 'alley', 'curtains', 'cliff', 'glass', 'finger', 'bandage'],
        'negative': ['ambulance', 'pain', 'disloyal', 'disaster', 'neglect', 'alone', 'betrayal'],
    },
    2: {
        'positive': ['diploma', 'joking', 'fun', 'rainbow', 'baby', 'romantic', 'cash'],
        'neutral': ['village', 'detail', 'cabinet', 'yellow', 'bench', 'errand', 'elevator'],
        'negative': ['agony', 'afraid', 'infection', 'abuse', 'death', 'bankrupt', 'cruel'],
    },
    3: {
        'positive': ['delight', 'friendly', 'humour', 'beach', 'free', 'joyful', 'comfort'],
        'neutral': ['ankle', 'headlight', 'machine', 'basket', 'grass', 'bowl', 'doctor'],
        'negative': ['offend', 'danger', 'stress', 'anger', 'funeral', 'evil', 'ugly'],
    },
}

# ============================================
# COLUMN MAPPING (old name -> new name)
# ============================================

COLUMNS_TO_KEEP = {
    'ResponseId': 'ResponseId',
    'Q5': 'Age',
    'Q6': 'Gender',
    'Q7': 'Sex',
    'Q8': 'Ethnicity',
    'Q9': 'Ethnicity_Other',
    'Q10': 'NativeEnglish',
    'Condition': 'Condition',
    'DominantFocus': 'DominantFocus',
    'RFQ1': 'RFQ1',
    'RFQ2': 'RFQ2',
    'RFQ3': 'RFQ3',
    'RFQ4': 'RFQ4',
    'RFQ5': 'RFQ5',
    'RFQ6': 'RFQ6',
    'RFQ7': 'RFQ7',
    'RFQ8': 'RFQ8',
    'RFQ9': 'RFQ9',
    'RFQ10': 'RFQ10',
    'RFQ11': 'RFQ11',
    'PrevScore': 'PrevScore',
    'PromoScore': 'PromoScore',
    'PrevSum': 'PrevSum',
    'PromoSum': 'PromoSum',
    # Distractor task answers (Block 1)
    'Q49': 'Distractor1_Q1',
    'Q50': 'Distractor1_Q2',
    'Q51': 'Distractor1_Q3',
    'Q52': 'Distractor1_Q4',
    'Q53': 'Distractor1_Q5',
    'Q54': 'Distractor1_Q6',
    'Q55': 'Distractor1_Q7',
    # Distractor task answers (Block 2)
    'Q78': 'Distractor2_Q1',
    'Q79': 'Distractor2_Q2',
    'Q80': 'Distractor2_Q3',
    'Q81': 'Distractor2_Q4',
    'Q82': 'Distractor2_Q5',
    'Q83': 'Distractor2_Q6',
    'Q84': 'Distractor2_Q7',
    # Distractor task answers (Block 3)
    'Q107': 'Distractor3_Q1',
    'Q108': 'Distractor3_Q2',
    'Q109': 'Distractor3_Q3',
    'Q110': 'Distractor3_Q4',
    'Q111': 'Distractor3_Q5',
    'Q112': 'Distractor3_Q6',
    'Q113': 'Distractor3_Q7',
    # Free recall responses (raw text)
    'Q482': 'Recall_Block1_Raw',
    'Q1460': 'Recall_Block2_Raw',
    'Q1462': 'Recall_Block3_Raw',
}

# Columns to output in order
OUTPUT_ORDER = [
    'ResponseId',
    'CompletionStatus',  # COMPLETE or INCOMPLETE
    'Age',
    'Gender',
    'Sex',
    'Ethnicity',
    'Ethnicity_Other',
    'NativeEnglish',
    'Condition',
    'DominantFocus',
    'RFQ1', 'RFQ2', 'RFQ3', 'RFQ4', 'RFQ5', 'RFQ6', 'RFQ7', 'RFQ8', 'RFQ9', 'RFQ10', 'RFQ11',
    'PrevScore', 'PromoScore', 'PrevSum', 'PromoSum',
    'Distractor1_Q1', 'Distractor1_Q2', 'Distractor1_Q3', 'Distractor1_Q4', 'Distractor1_Q5', 'Distractor1_Q6', 'Distractor1_Q7',
    'Distractor2_Q1', 'Distractor2_Q2', 'Distractor2_Q3', 'Distractor2_Q4', 'Distractor2_Q5', 'Distractor2_Q6', 'Distractor2_Q7',
    'Distractor3_Q1', 'Distractor3_Q2', 'Distractor3_Q3', 'Distractor3_Q4', 'Distractor3_Q5', 'Distractor3_Q6', 'Distractor3_Q7',
    'Recall_Block1_Raw', 'Recall_Block2_Raw', 'Recall_Block3_Raw',
    # Scored columns (added by script)
    'Recall_Block1_Total', 'Recall_Block1_Pos', 'Recall_Block1_Neu', 'Recall_Block1_Neg',
    'Recall_Block2_Total', 'Recall_Block2_Pos', 'Recall_Block2_Neu', 'Recall_Block2_Neg',
    'Recall_Block3_Total', 'Recall_Block3_Pos', 'Recall_Block3_Neu', 'Recall_Block3_Neg',
    'Recall_Total', 'Recall_Total_Pos', 'Recall_Total_Neu', 'Recall_Total_Neg',
]

# ============================================
# FUZZY MATCHING & SCORING
# ============================================

def normalize_word(word):
    """Lowercase, strip, remove punctuation."""
    word = word.lower().strip()
    word = re.sub(r'[^a-z]', '', word)
    return word

# UK/US spelling equivalents
UK_US_SPELLINGS = {
    'humor': 'humour',
    'color': 'colour',
    'favor': 'favour',
    'neighbor': 'neighbour',
}

def normalize_spelling(word):
    """Convert US spellings to UK equivalents."""
    return UK_US_SPELLINGS.get(word, word)

def simple_deplural(word):
    """Only handle simple plurals (adding 's'), not word transformations."""
    if word.endswith('s') and len(word) > 3 and not word.endswith(('ss', 'us', 'is')):
        return word[:-1]
    return word

def fuzzy_match(input_word, target_word, threshold=0.92):
    """
    Check if input_word matches target_word with fuzzy matching.
    Threshold raised to 0.92 to prevent false positives like friendship/friendly.
    Only allows for minor typos (1-2 chars) not word variations.
    """
    input_norm = normalize_word(input_word)
    target_norm = normalize_word(target_word)
    
    # Normalize UK/US spellings
    input_norm = normalize_spelling(input_norm)
    target_norm = normalize_spelling(target_norm)
    
    # Exact match
    if input_norm == target_norm:
        return True
    
    # Simple plural match (bandages -> bandage)
    if simple_deplural(input_norm) == target_norm:
        return True
    if input_norm == simple_deplural(target_norm):
        return True
    
    # Fuzzy match only for very close matches (typos)
    # Must be similar length (within 2 chars) to prevent friendship/friendly
    if abs(len(input_norm) - len(target_norm)) > 2:
        return False
    
    ratio = SequenceMatcher(None, input_norm, target_norm).ratio()
    if ratio >= threshold:
        return True
    
    return False

def split_camelcase(text):
    """
    Split CamelCase or concatenated words into separate words.
    Handles cases like 'BandageDiplomaErrand' -> ['Bandage', 'Diploma', 'Errand']
    Also handles 'bandagediplomaerrand' by matching against known words.
    """
    # First try splitting on capital letters (CamelCase)
    # This regex finds positions where a lowercase is followed by uppercase
    camel_split = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)
    if camel_split != text:
        return camel_split.split()
    
    # If no camelCase detected, try to find known words in the string
    # Build a list of all known words
    all_words = []
    for block in WORDS.values():
        for valence_words in block.values():
            all_words.extend(valence_words)
    
    # Sort by length (longest first) to match greedily
    all_words = sorted(set(all_words), key=len, reverse=True)
    
    text_lower = text.lower()
    found_words = []
    remaining = text_lower
    
    while remaining:
        matched = False
        for word in all_words:
            if remaining.startswith(word):
                found_words.append(word)
                remaining = remaining[len(word):]
                matched = True
                break
        if not matched:
            # No known word found at start, skip one character
            remaining = remaining[1:]
    
    return found_words if found_words else [text]

def parse_recall_response(text):
    """Split recall response into individual words."""
    if not text or not text.strip():
        return []
    
    # Split on commas, spaces, newlines, or multiple spaces
    words = re.split(r'[,\s\n]+', text)
    words = [w.strip() for w in words if w.strip()]
    
    # Process each word for potential CamelCase or concatenated words
    expanded_words = []
    for word in words:
        # If word is unusually long (likely concatenated), try to split it
        if len(word) > 12:
            split_words = split_camelcase(word)
            expanded_words.extend(split_words)
        else:
            expanded_words.append(word)
    
    return expanded_words

def score_recall(response_text, block_num):
    """
    Score a recall response for a given block.
    Returns dict with total, positive, neutral, negative counts.
    """
    recalled_words = parse_recall_response(response_text)
    
    scores = {'total': 0, 'positive': 0, 'neutral': 0, 'negative': 0}
    matched_targets = set()  # Prevent double-counting
    
    for recalled in recalled_words:
        for valence in ['positive', 'neutral', 'negative']:
            for target in WORDS[block_num][valence]:
                if target not in matched_targets and fuzzy_match(recalled, target):
                    scores['total'] += 1
                    scores[valence] += 1
                    matched_targets.add(target)
                    break
            else:
                continue
            break
    
    return scores

# ============================================
# EXCEL OUTPUT WITH COLOR CODING
# ============================================

# Pastel color scheme (easy on eyes)
COLORS = {
    'id': 'E8E8E8',          # Light grey - ResponseId
    'demographics': 'D4E6F1', # Soft blue - Age, Gender, etc.
    'ivs': 'D5F5E3',          # Soft green - Condition, DominantFocus
    'rfq': 'FCF3CF',          # Soft yellow - RFQ items
    'rfq_scores': 'F9E79F',   # Darker yellow - RFQ computed scores
    'distractor': 'FADBD8',   # Soft pink - Distractor tasks
    'recall_raw': 'E8DAEF',   # Soft purple - Raw recall text
    'scores': 'ABEBC6',       # Soft green - Computed recall scores
    'header': 'ABB2B9',       # Grey - Header row
}

# Column categories
COLUMN_CATEGORIES = {
    'ResponseId': 'id',
    'CompletionStatus': 'ivs',
    'Age': 'demographics',
    'Gender': 'demographics',
    'Sex': 'demographics',
    'Ethnicity': 'demographics',
    'Ethnicity_Other': 'demographics',
    'NativeEnglish': 'demographics',
    'Condition': 'ivs',
    'DominantFocus': 'ivs',
    'PrevScore': 'rfq_scores',
    'PromoScore': 'rfq_scores',
    'PrevSum': 'rfq_scores',
    'PromoSum': 'rfq_scores',
    'Recall_Block1_Raw': 'recall_raw',
    'Recall_Block2_Raw': 'recall_raw',
    'Recall_Block3_Raw': 'recall_raw',
}

def get_column_category(col_name):
    if col_name in COLUMN_CATEGORIES:
        return COLUMN_CATEGORIES[col_name]
    if col_name.startswith('RFQ'):
        return 'rfq'
    if col_name.startswith('Distractor'):
        return 'distractor'
    if col_name.startswith('Recall_') and '_Raw' not in col_name:
        return 'scores'
    return 'id'  # Default

def write_excel(output_rows, output_path):
    """Write output as formatted Excel file."""
    if not EXCEL_SUPPORT:
        print("Error: openpyxl not installed. Install with: pip install openpyxl")
        return False
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    
    # Create fills for each category
    fills = {cat: PatternFill(start_color=color, end_color=color, fill_type='solid') 
             for cat, color in COLORS.items()}
    
    # Header style
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Write header row
    for col_idx, col_name in enumerate(OUTPUT_ORDER, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fills['header']
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Write data rows
    for row_idx, row_data in enumerate(output_rows, 2):
        for col_idx, col_name in enumerate(OUTPUT_ORDER, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            category = get_column_category(col_name)
            cell.fill = fills[category]
            cell.border = thin_border
            
            # Center align numbers
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(OUTPUT_ORDER, 1):
        max_length = len(col_name)
        for row_idx in range(2, len(output_rows) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_length = max(max_length, min(len(cell_value), 50))  # Cap at 50
        
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    return True

# ============================================
# MANUAL CONDITION OVERRIDES (test entries with missing conditions)
# ============================================

CONDITION_OVERRIDES = {
    'R_83vAa699ie85a48': 'prevention',
    'R_8PaPZ55rpsVFSyZ': 'promotion',
}

# ============================================
# MAIN PROCESSING
# ============================================

def process_csv(input_path, output_path):
    with open(input_path, 'r', encoding='utf-8-sig') as infile:
        reader = csv.DictReader(infile)
        rows = list(reader)
        
        # Filter out non-data rows (Qualtrics metadata)
        data_rows = []
        for row in rows:
            response_id = row.get('ResponseId', '')
            if response_id.startswith('Response') or response_id.startswith('Start Date'):
                continue
            if 'ImportId' in response_id:
                continue
            if response_id.startswith('R_'):
                # Apply manual condition overrides
                if response_id in CONDITION_OVERRIDES:
                    row['Condition'] = CONDITION_OVERRIDES[response_id]
                data_rows.append(row)
        
        rows = data_rows
        output_rows = []
        
        for row in rows:
            new_row = {}
            
            # Copy and rename columns
            for old_col, new_col in COLUMNS_TO_KEEP.items():
                new_row[new_col] = row.get(old_col, '')
            
            # Score each block
            block1_scores = score_recall(new_row.get('Recall_Block1_Raw', ''), 1)
            block2_scores = score_recall(new_row.get('Recall_Block2_Raw', ''), 2)
            block3_scores = score_recall(new_row.get('Recall_Block3_Raw', ''), 3)
            
            # Add block scores
            new_row['Recall_Block1_Total'] = block1_scores['total']
            new_row['Recall_Block1_Pos'] = block1_scores['positive']
            new_row['Recall_Block1_Neu'] = block1_scores['neutral']
            new_row['Recall_Block1_Neg'] = block1_scores['negative']
            
            new_row['Recall_Block2_Total'] = block2_scores['total']
            new_row['Recall_Block2_Pos'] = block2_scores['positive']
            new_row['Recall_Block2_Neu'] = block2_scores['neutral']
            new_row['Recall_Block2_Neg'] = block2_scores['negative']
            
            new_row['Recall_Block3_Total'] = block3_scores['total']
            new_row['Recall_Block3_Pos'] = block3_scores['positive']
            new_row['Recall_Block3_Neu'] = block3_scores['neutral']
            new_row['Recall_Block3_Neg'] = block3_scores['negative']
            
            # Add totals
            new_row['Recall_Total'] = block1_scores['total'] + block2_scores['total'] + block3_scores['total']
            new_row['Recall_Total_Pos'] = block1_scores['positive'] + block2_scores['positive'] + block3_scores['positive']
            new_row['Recall_Total_Neu'] = block1_scores['neutral'] + block2_scores['neutral'] + block3_scores['neutral']
            new_row['Recall_Total_Neg'] = block1_scores['negative'] + block2_scores['negative'] + block3_scores['negative']
            
            # Determine completion status
            # COMPLETE = ALL THREE blocks must have at least one correctly recalled word
            # INCOMPLETE = any block has zero recalled words (participant didn't complete all blocks)
            # Uses actual scored totals, not just raw text presence
            
            if (new_row['Recall_Block1_Total'] >= 1 and 
                new_row['Recall_Block2_Total'] >= 1 and 
                new_row['Recall_Block3_Total'] >= 1):
                new_row['CompletionStatus'] = 'COMPLETE'
            else:
                new_row['CompletionStatus'] = 'INCOMPLETE'
            
            output_rows.append(new_row)
        
        # Determine output format
        if output_path.endswith('.xlsx'):
            if write_excel(output_rows, output_path):
                print(f"✓ Processed {len(output_rows)} responses")
                print(f"✓ Excel file saved to: {output_path}")
            else:
                print("Failed to write Excel file")
                sys.exit(1)
        else:
            # Write CSV
            with open(output_path, 'w', encoding='utf-8', newline='') as outfile:
                writer = csv.DictWriter(outfile, fieldnames=OUTPUT_ORDER)
                writer.writeheader()
                writer.writerows(output_rows)
            
            print(f"✓ Processed {len(output_rows)} responses")
            print(f"✓ CSV saved to: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python3 qualtrics_cleanup.py input.csv output.csv")
        print("       python3 qualtrics_cleanup.py input.csv output.xlsx  (colored Excel)")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    process_csv(input_file, output_file)
